import openpyxl
import pymysql.cursors

# Excel file
wb = openpyxl.load_workbook('./movies.xlsx')
# Get the active worksheet
ws = wb.active

# Add columns to the map function and increment x[] accordingly
#ws[1:] is the row to start at
data = map(lambda x: {'name': x[0].value,
                      'count': x[1].value,
                      'score': x[2].value,
                      'content': x[3].value
                      },

           ws[1: ws.max_row])


# Checks for empty cells
data = filter(lambda x: None not in x.values(), data)
print(data)
db = pymysql.connect(host='172.18.18.203', port=23306, user='root', password='root', db='movies', charset='utf8mb4', cursorclass=pymysql.cursors.DictCursor)

cursor = db.cursor()

for row in data:
    cursor.execute('INSERT INTO `movies` (`name`,`count`,`score`,`content`)'
                   'VALUES ("{name}","{count}","{score}","{content}");'
                   .format(**row))  # construct MySQL syntax through format function
    print(row)
db.commit()